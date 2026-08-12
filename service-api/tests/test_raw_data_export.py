from datetime import date, datetime, timezone
from io import BytesIO

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError

from cubici_service.preferences.repository import RawDataExportRequest, build_raw_data_workbook


def test_raw_data_export_request_rejects_partial_or_long_date_range() -> None:
    with pytest.raises(ValidationError):
        RawDataExportRequest(table_name="sale", columns=["order_no"], from_date=date(2026, 1, 1))

    with pytest.raises(ValidationError):
        RawDataExportRequest(
            table_name="sale",
            columns=["order_no"],
            from_date=date(2025, 1, 1),
            to_date=date(2026, 2, 1),
        )


def test_raw_data_workbook_sanitizes_formula_like_values() -> None:
    content = build_raw_data_workbook(
        ["order_no", "amount"],
        [{"order_no": "=HYPERLINK(\"https://example.invalid\")", "amount": 1000}],
    )

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    worksheet = workbook["RawData"]
    assert worksheet["A1"].value == "order_no"
    assert worksheet["A2"].value == "'=HYPERLINK(\"https://example.invalid\")"
    assert worksheet["B2"].value == 1000


def test_raw_data_workbook_normalizes_json_and_timezone_values() -> None:
    content = build_raw_data_workbook(
        ["metadata", "paid_at"],
        [{"metadata": {"status": "정상"}, "paid_at": datetime(2026, 8, 11, 9, 30, tzinfo=timezone.utc)}],
    )

    workbook = load_workbook(BytesIO(content), read_only=True)
    worksheet = workbook["RawData"]
    assert worksheet["A2"].value == '{"status": "정상"}'
    assert worksheet["B2"].value == datetime(2026, 8, 11, 9, 30)
